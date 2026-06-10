from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, jsonify
from flask_login import login_required, current_user
from sqlalchemy import func, and_
from models import db, Branch, Department, Item, InventoryCount, SessionAuditLog
from datetime import datetime, date
import calendar
from utils.constants import now_jordan
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

analytics_bp = Blueprint('analytics', __name__, url_prefix='/analytics')

_ZERO_THRESHOLD = 0.001   # floating-point tolerance for "no change"
_OLIVE  = '4B5947'
_GREEN  = '295831'
_ORANGE = 'F28705'


def _fmt(n):
    """Format a float: integer if whole, else up to 3 decimal places."""
    if not n:
        return '0'
    n = float(n)
    if n == int(n):
        return str(int(n))
    return f'{n:.3f}'.rstrip('0').rstrip('.')


# ── Filter helper ─────────────────────────────────────────────────────────────

def _parse_filters():
    now        = now_jordan()
    branch_id  = request.args.get('branch_id',  type=int)
    dept_id    = request.args.get('dept_id',    type=int)
    session_id = request.args.get('session_id', type=int)
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to',   '')

    date_from = date_to = None
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    except ValueError:
        pass
    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        pass

    if not date_from and not date_to:
        date_from = date(now.year, now.month, 1)
        date_to   = date(now.year, now.month, calendar.monthrange(now.year, now.month)[1])
        date_from_str = date_from.isoformat()
        date_to_str   = date_to.isoformat()

    return branch_id, dept_id, date_from, date_to, date_from_str, date_to_str, session_id


# ── Core analytics SQL ─────────────────────────────────────────────────────────

def _analytics_data(branch_id, dept_id, date_from, date_to, session_id=None):
    """
    Five-subquery approach — all heavy work in SQL, single round-trip.

    Primary filter: session_id (when provided — session is the grouping key).
    Fallback filter: count_date range (backward-compat date-based filtering).

    Returns list of tuples:
        (Item, Department, Branch,
         entry_count, total_qty, min_qty, max_qty,
         first_date, last_date, first_qty, last_qty)
    """

    # ── Subquery 1: per-item aggregation (counts, total, date bounds) ─────────
    agg_q = (
        db.session.query(
            InventoryCount.item_id.label('item_id'),
            func.count(InventoryCount.id).label('entry_count'),
            func.sum(InventoryCount.quantity).label('total_qty'),
            func.min(InventoryCount.count_date).label('first_date'),
            func.max(InventoryCount.count_date).label('last_date'),
        )
        .join(Item,       InventoryCount.item_id      == Item.id)
        .join(Department, Item.department_id           == Department.id)
    )
    if session_id:
        agg_q = agg_q.filter(InventoryCount.session_id == session_id)
    else:
        agg_q = agg_q.filter(
            InventoryCount.count_date >= date_from,
            InventoryCount.count_date <= date_to,
        )
    if branch_id:
        agg_q = agg_q.filter(Department.branch_id == branch_id)
    if dept_id:
        agg_q = agg_q.filter(Item.department_id == dept_id)
    agg_q = agg_q.filter(InventoryCount.status == 'active')
    agg_q = agg_q.group_by(InventoryCount.item_id).subquery('agg')

    # ── Subquery 2: total qty on first_date per item ──────────────────────────
    fq = (
        db.session.query(
            InventoryCount.item_id.label('item_id'),
            func.sum(InventoryCount.quantity).label('first_qty'),
        )
        .join(agg_q, and_(
            InventoryCount.item_id    == agg_q.c.item_id,
            InventoryCount.count_date == agg_q.c.first_date,
        ))
        .filter(InventoryCount.status == 'active')
        .group_by(InventoryCount.item_id)
        .subquery('fq')
    )

    # ── Subquery 3: total qty on last_date per item ───────────────────────────
    lq = (
        db.session.query(
            InventoryCount.item_id.label('item_id'),
            func.sum(InventoryCount.quantity).label('last_qty'),
        )
        .join(agg_q, and_(
            InventoryCount.item_id    == agg_q.c.item_id,
            InventoryCount.count_date == agg_q.c.last_date,
        ))
        .filter(InventoryCount.status == 'active')
        .group_by(InventoryCount.item_id)
        .subquery('lq')
    )

    # ── Subquery 4: daily totals per (item, date) ─────────────────────────────
    daily_q = (
        db.session.query(
            InventoryCount.item_id.label('item_id'),
            InventoryCount.count_date.label('count_date'),
            func.sum(InventoryCount.quantity).label('day_total'),
        )
        .join(Item,       InventoryCount.item_id  == Item.id)
        .join(Department, Item.department_id       == Department.id)
    )
    if session_id:
        daily_q = daily_q.filter(InventoryCount.session_id == session_id)
    else:
        daily_q = daily_q.filter(
            InventoryCount.count_date >= date_from,
            InventoryCount.count_date <= date_to,
        )
    if branch_id:
        daily_q = daily_q.filter(Department.branch_id == branch_id)
    if dept_id:
        daily_q = daily_q.filter(Item.department_id == dept_id)
    daily_q = daily_q.filter(InventoryCount.status == 'active')
    daily_q = daily_q.group_by(
        InventoryCount.item_id, InventoryCount.count_date
    ).subquery('daily')

    # ── Subquery 5: min/max of daily totals per item ──────────────────────────
    minmax_q = (
        db.session.query(
            daily_q.c.item_id.label('item_id'),
            func.min(daily_q.c.day_total).label('min_qty'),
            func.max(daily_q.c.day_total).label('max_qty'),
        )
        .group_by(daily_q.c.item_id)
        .subquery('minmax')
    )

    # ── Final join: Item + Department + Branch + all stats ────────────────────
    return (
        db.session.query(
            Item,
            Department,
            Branch,
            agg_q.c.entry_count,
            agg_q.c.total_qty,
            minmax_q.c.min_qty,
            minmax_q.c.max_qty,
            agg_q.c.first_date,
            agg_q.c.last_date,
            fq.c.first_qty,
            lq.c.last_qty,
        )
        .join(agg_q,      Item.id               == agg_q.c.item_id)
        .join(fq,         Item.id               == fq.c.item_id)
        .join(lq,         Item.id               == lq.c.item_id)
        .join(minmax_q,   Item.id               == minmax_q.c.item_id)
        .join(Department, Item.department_id     == Department.id)
        .join(Branch,     Department.branch_id   == Branch.id)
        .order_by(Branch.name_ar, Department.name_ar, Item.name_ar)
        .all()
    )


# ── Row processing ─────────────────────────────────────────────────────────────

def _process_rows(raw_rows):
    """Convert raw DB tuples into grouped dicts + summary counters."""
    groups      = {}
    group_order = []
    increases = decreases = unchanged = 0
    needs_review_count = modified_count = 0

    for (item, dept, branch,
         entry_count, total_qty, min_qty, max_qty,
         first_date, last_date, first_qty, last_qty) in raw_rows:

        first_qty  = first_qty or 0.0
        last_qty   = last_qty  or 0.0
        net_change = last_qty - first_qty

        if net_change > _ZERO_THRESHOLD:
            increases += 1
            trend = 'up'
        elif net_change < -_ZERO_THRESHOLD:
            decreases += 1
            trend = 'down'
        else:
            unchanged += 1
            trend = 'flat'

        key = f'{branch.id}_{dept.id}'
        if key not in groups:
            groups[key] = {
                'branch_name': branch.name_ar,
                'dept_name':   dept.name_ar,
                'rows':        [],
            }
            group_order.append(key)

        ms       = item.effective_minimum_stock
        low_stock = ms > 0 and last_qty < ms

        if low_stock:
            status = 'needs_review'
            needs_review_count += 1
        elif entry_count > 1:
            status = 'modified'
            modified_count += 1
        else:
            status = 'counted'

        groups[key]['rows'].append({
            'item':           item,
            'entry_count':    entry_count,
            'total_qty':      total_qty      or 0.0,
            'min_qty':        min_qty        or 0.0,
            'max_qty':        max_qty        or 0.0,
            'first_date':     first_date,
            'last_date':      last_date,
            'first_qty':      first_qty,
            'last_qty':       last_qty,
            'net_change':     net_change,
            'trend':          trend,
            'minimum_stock':  ms,
            'low_stock':      low_stock,
            'status':         status,
        })

    return [groups[k] for k in group_order], increases, decreases, unchanged, needs_review_count, modified_count


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(groups, increases, decreases, unchanged, total_items,
                 date_from_str, date_to_str, branch_label, dept_label):
    # (header_text, col_width, cell_alignment)
    COLS = [
        ('#',                      5,  'center'),
        ('كود الصنف',             16,  'center'),
        ('اسم الصنف',             35,  'right'),
        ('الوحدة',                12,  'center'),
        ('الكمية الأولى',         14,  'center'),
        ('الكمية الأخيرة',        14,  'center'),
        ('الفرق',                 12,  'center'),
        ('الحد الأدنى للمخزون',   15,  'center'),  # col 8 — minimum_stock threshold
        ('أقل يومي',              13,  'center'),  # col 9 — min recorded daily total
        ('أعلى يومي',             13,  'center'),  # col 10 — max recorded daily total
        ('عدد الإدخالات',         16,  'center'),
        ('تاريخ أول جرد',         15,  'center'),
        ('تاريخ آخر جرد',         15,  'center'),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تحليل الجرد'
    ws.sheet_view.rightToLeft = True

    # Set column widths up front
    for ci, (_, w, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, val='', bold=False, size=10,
             fg='000000', bg=None, ha='center', wrap=False, border=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, size=size, color=fg)
        c.alignment = Alignment(horizontal=ha, vertical='center',
                                readingOrder=2, wrap_text=wrap)
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        if border:
            c.border = brd
        return c

    r = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:M{r}')
    cell(r, 1, 'مطاعم همم — تحليل حركة الجرد',
         bold=True, size=16, fg=_OLIVE, ha='center')
    ws.row_dimensions[r].height = 38
    r += 1

    # ── Period / filters ──────────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:M{r}')
    period = f'الفترة: {date_from_str} — {date_to_str}'
    if branch_label:
        period += f'  |  الفرع: {branch_label}'
    if dept_label:
        period += f'  |  القسم: {dept_label}'
    cell(r, 1, period, size=11, fg='555555', ha='center')
    ws.row_dimensions[r].height = 22
    r += 1

    ws.merge_cells(f'A{r}:M{r}')
    cell(r, 1, f'تاريخ الإصدار: {date.today().isoformat()}',
         size=9, fg='888888', ha='center')
    ws.row_dimensions[r].height = 18
    r += 1

    r += 1  # blank row

    # ── Summary stats (2 rows × 4 blocks A:B, C:D, E:F, G:H) ────────────────
    stats = [
        ('A', 'B', 'إجمالي الأصناف', total_items,  _OLIVE),
        ('C', 'D', 'ارتفعت كميتها',  increases,   '198754'),
        ('E', 'F', 'انخفضت كميتها', decreases,   'dc3545'),
        ('G', 'H', 'بدون تغيير',    unchanged,   '6c757d'),
    ]
    for ca, cb, label, val, color in stats:
        ws.merge_cells(f'{ca}{r}:{cb}{r}')
        ws.merge_cells(f'{ca}{r+1}:{cb}{r+1}')
        ci = ord(ca) - 64
        cell(r,   ci, label, bold=True, size=9,  fg='FFFFFF', bg=color, ha='center')
        cell(r+1, ci, val,   bold=True, size=14, fg=color,              ha='center')
    ws.row_dimensions[r].height   = 22
    ws.row_dimensions[r+1].height = 28
    r += 3   # 2 stat rows + 1 blank

    r += 1   # second blank before table header

    # ── Table header + freeze ──────────────────────────────────────────────────
    for ci, (label, _, _) in enumerate(COLS, 1):
        cell(r, ci, label, bold=True, size=10, fg='FFFFFF',
             bg=_OLIVE, ha='center', wrap=True, border=True)
    ws.row_dimensions[r].height = 32
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    r += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    for group in groups:
        # Department / branch group header
        ws.merge_cells(f'A{r}:M{r}')
        cell(r, 1, f"{group['dept_name']}  —  {group['branch_name']}",
             bold=True, size=10, fg='FFFFFF', bg=_GREEN, ha='right')
        ws.row_dimensions[r].height = 22
        r += 1

        dept_first = dept_last = 0.0

        for ri, rd in enumerate(group['rows'], 1):
            item = rd['item']
            net  = rd['net_change']
            low  = rd['low_stock']
            tint = ('fff3cd' if low else
                    'eaf6ee' if rd['trend'] == 'up' else
                    'fef0ef' if rd['trend'] == 'down' else None)

            row_vals = [
                ri,
                item.item_code or '',
                item.name_ar,
                item.effective_base_unit.name_ar,
                rd['first_qty'],
                rd['last_qty'],
                net,
                rd['minimum_stock'],   # col 8 — threshold
                rd['min_qty'],         # col 9 — lowest daily recorded
                rd['max_qty'],         # col 10 — highest daily recorded
                rd['entry_count'],
                rd['first_date'].isoformat() if rd['first_date'] else '',
                rd['last_date'].isoformat() if rd['last_date'] else '',
            ]

            for ci, (val, (_, _, align)) in enumerate(zip(row_vals, COLS), 1):
                c = cell(r, ci, val, size=9, ha=align, bg=tint, border=True)
                if ci == 7:  # net change — colour-code the font
                    if net > _ZERO_THRESHOLD:
                        c.font = Font(bold=True, size=9, color='198754')
                    elif net < -_ZERO_THRESHOLD:
                        c.font = Font(bold=True, size=9, color='dc3545')
                    else:
                        c.font = Font(size=9, color='888888')
                    c.alignment = Alignment(horizontal='center', readingOrder=2)
                elif ci == 8 and low:  # minimum_stock cell — highlight when low
                    c.font = Font(bold=True, size=9, color='dc3545')
                    c.fill = PatternFill('solid', fgColor='ffebee')
                    c.alignment = Alignment(horizontal='center', readingOrder=2)

            dept_first += rd['first_qty']
            dept_last  += rd['last_qty']
            r += 1

        # Department totals footer
        dept_change = dept_last - dept_first
        for ci in range(1, 14):
            c = ws.cell(row=r, column=ci, value='')
            c.fill   = PatternFill('solid', fgColor='e0e9e0')
            c.font   = Font(bold=True, size=9, color=_OLIVE)
            c.alignment = Alignment(horizontal='center', readingOrder=2)
            c.border = brd

        ws.cell(row=r, column=3).value     = f"إجمالي: {group['dept_name']}"
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='right', readingOrder=2)
        ws.cell(row=r, column=5).value = dept_first
        ws.cell(row=r, column=6).value = dept_last
        ws.cell(row=r, column=7).value = dept_change
        nc = ws.cell(row=r, column=7)
        if dept_change > _ZERO_THRESHOLD:
            nc.font = Font(bold=True, size=9, color='198754')
        elif dept_change < -_ZERO_THRESHOLD:
            nc.font = Font(bold=True, size=9, color='dc3545')
        ws.row_dimensions[r].height = 20
        r += 1
        r += 1   # blank row between groups

    return wb


# ── Routes ────────────────────────────────────────────────────────────────────

@analytics_bp.route('/')
@login_required
def index():
    if not current_user.is_manager:
        flash('هذه الصفحة للمدراء فقط', 'warning')
        return redirect(url_for('inventory.dashboard'))

    branch_id, dept_id, date_from, date_to, date_from_str, date_to_str, session_id = _parse_filters()

    # Non-admin managers see only their own branch
    if not current_user.is_admin and current_user.branch_id:
        branch_id = current_user.branch_id

    branches    = Branch.query.order_by(Branch.name_ar).all()
    departments = (
        Department.query
        .filter_by(branch_id=branch_id).order_by(Department.name_ar).all()
        if branch_id else
        Department.query.order_by(Department.name_ar).all()
    )

    raw_rows = _analytics_data(branch_id, dept_id, date_from, date_to, session_id)
    groups, increases, decreases, unchanged, needs_review_count, modified_count = _process_rows(raw_rows)

    counted_items = increases + decreases + unchanged
    changes_count = increases + decreases
    counted_clean = counted_items - needs_review_count - modified_count

    # Total active items in scope (for uncounted KPI)
    scope_q = Item.query.filter_by(is_active=True)
    if branch_id:
        scope_q = scope_q.join(Department, Item.department_id == Department.id).filter(
            Department.branch_id == branch_id
        )
    if dept_id:
        scope_q = scope_q.filter(Item.department_id == dept_id)
    total_in_scope  = scope_q.count()
    uncounted_items = max(0, total_in_scope - counted_items)

    return render_template(
        'inventory/analytics.html',
        groups=groups,
        total_items=counted_items,
        total_in_scope=total_in_scope,
        uncounted_items=uncounted_items,
        increases=increases,
        decreases=decreases,
        unchanged=unchanged,
        changes_count=changes_count,
        needs_review_count=needs_review_count,
        modified_count=modified_count,
        counted_clean=counted_clean,
        branches=branches,
        departments=departments,
        selected_branch=branch_id,
        selected_dept=dept_id,
        selected_session=session_id,
        date_from=date_from_str,
        date_to=date_to_str,
        now=now_jordan(),
    )


@analytics_bp.route('/export/excel')
@login_required
def export_excel():
    if not current_user.is_manager:
        flash('التصدير للمدراء فقط', 'warning')
        return redirect(url_for('analytics.index'))

    branch_id, dept_id, date_from, date_to, date_from_str, date_to_str, session_id = _parse_filters()
    if not current_user.is_admin and current_user.branch_id:
        branch_id = current_user.branch_id

    raw_rows = _analytics_data(branch_id, dept_id, date_from, date_to, session_id)
    groups, increases, decreases, unchanged, *_ = _process_rows(raw_rows)
    total_items = increases + decreases + unchanged

    branch_label = dept_label = ''
    if branch_id:
        b = Branch.query.get(branch_id)
        if b:
            branch_label = b.name_ar
    if dept_id:
        d = Department.query.get(dept_id)
        if d:
            dept_label = d.name_ar

    wb  = _build_excel(groups, increases, decreases, unchanged, total_items,
                       date_from_str, date_to_str, branch_label, dept_label)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventory_analytics_{date_from_str}_{date_to_str}.xlsx',
    )


@analytics_bp.route('/export/pdf')
@login_required
def export_pdf():
    if not current_user.is_manager:
        flash('التصدير للمدراء فقط', 'warning')
        return redirect(url_for('analytics.index'))

    branch_id, dept_id, date_from, date_to, date_from_str, date_to_str, session_id = _parse_filters()
    if not current_user.is_admin and current_user.branch_id:
        branch_id = current_user.branch_id

    raw_rows = _analytics_data(branch_id, dept_id, date_from, date_to, session_id)
    groups, increases, decreases, unchanged, *_ = _process_rows(raw_rows)

    branch_label = dept_label = ''
    if branch_id:
        b = Branch.query.get(branch_id)
        if b:
            branch_label = b.name_ar
    if dept_id:
        d = Department.query.get(dept_id)
        if d:
            dept_label = d.name_ar

    return render_template(
        'inventory/analytics_pdf.html',
        groups=groups,
        total_items=increases + decreases + unchanged,
        increases=increases,
        decreases=decreases,
        unchanged=unchanged,
        date_from=date_from_str,
        date_to=date_to_str,
        branch_label=branch_label,
        dept_label=dept_label,
        now=now_jordan(),
    )


@analytics_bp.route('/item/<int:item_id>/history')
@login_required
def item_history(item_id):
    if not current_user.is_manager:
        return jsonify({'error': 'unauthorized'}), 403

    item = Item.query.get_or_404(item_id)
    session_id    = request.args.get('session_id',  type=int)
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to',   '')

    date_from = date_to = None
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    except ValueError:
        pass
    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        pass

    q = InventoryCount.query.filter_by(item_id=item_id)
    if session_id:
        q = q.filter_by(session_id=session_id)
    else:
        if date_from:
            q = q.filter(InventoryCount.count_date >= date_from)
        if date_to:
            q = q.filter(InventoryCount.count_date <= date_to)
    entries = q.order_by(InventoryCount.created_at.asc()).all()

    al_q = SessionAuditLog.query.filter_by(item_id=item_id)
    if session_id:
        al_q = al_q.filter_by(session_id=session_id)
    elif date_from or date_to:
        if date_from:
            al_q = al_q.filter(SessionAuditLog.changed_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            al_q = al_q.filter(SessionAuditLog.changed_at <= datetime.combine(date_to, datetime.max.time()))
    audit_logs = al_q.order_by(SessionAuditLog.changed_at.asc()).all()

    unit_name = item.effective_base_unit.name_ar if item.effective_base_unit else ''

    return jsonify({
        'item': {
            'id':             item.id,
            'name_ar':        item.name_ar,
            'item_code':      item.item_code or '',
            'unit':           unit_name,
            'packaging_note': item.packaging_note or '',
            'min_stock':      float(item.effective_minimum_stock or 0),
        },
        'entries': [{
            'id':               e.id,
            'quantity':         float(e.quantity or 0),
            'entered_quantity': float(e.entered_quantity) if e.entered_quantity else None,
            'entered_unit':     e.entered_unit.name_ar if e.entered_unit else '',
            'user_name':        e.user.full_name if e.user else '—',
            'count_date':       e.count_date.isoformat() if e.count_date else '',
            'created_at':       e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
            'notes':            e.notes or '',
            'status':           getattr(e, 'status', 'active'),
        } for e in entries],
        'audit_logs': [{
            'changed_at':  al.changed_at.strftime('%Y-%m-%d %H:%M') if al.changed_at else '',
            'editor':      al.editor.full_name if al.editor else '—',
            'old_value':   al.old_value or '',
            'new_value':   al.new_value or '',
            'reason':      al.reason or '',
            'action_type': getattr(al, 'action_type', 'edit'),
        } for al in audit_logs],
    })
