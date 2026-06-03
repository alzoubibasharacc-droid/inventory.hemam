import csv
import io

from flask import render_template, request, make_response, jsonify
from models import db, Item, Unit, Branch, Department, Category, ItemUnitConversion
from utils.decorators import manager_required
from routes.admin import admin_bp


# ─── EXPORT ──────────────────────────────────────────────────────────────────

@admin_bp.route('/items/export')
@manager_required
def export_items():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    items = (
        Item.query
        .options(
            db.joinedload(Item.base_unit),
            db.joinedload(Item.category),
            db.joinedload(Item.department),
        )
        .order_by(Item.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'items'

    HEADERS = [
        'item_id', 'item_code', 'item_name_ar', 'item_name_en',
        'category_name', 'unit_name', 'min_level', 'notes',
        'is_active', 'created_at',
    ]

    header_fill = PatternFill('solid', fgColor='295831')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for item in items:
        base_unit = item.base_unit or item.unit
        ws.append([
            item.id,
            item.item_code or '',
            item.name_ar,
            item.name_en or '',
            item.category.name_ar if item.category else '',
            base_unit.name_ar if base_unit else '',
            item.effective_minimum_stock,
            item.packaging_note or '',
            '1' if item.is_active else '0',
            item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else '',
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp.headers['Content-Disposition'] = 'attachment; filename=items_export.xlsx'
    return resp


# ─── IMPORT ──────────────────────────────────────────────────────────────────

@admin_bp.route('/import', methods=['GET', 'POST'])
@manager_required
def import_items():
    branches     = Branch.query.all()
    departments  = Department.query.all()
    units        = Unit.query.order_by(Unit.name_ar).all()
    active_units = Unit.query.filter_by(is_active=True).order_by(Unit.name_ar).all()

    if request.method == 'GET':
        return render_template(
            'admin/import.html',
            branches=branches,
            departments=departments,
            units=units,
            active_units=active_units,
        )

    # ── POST ─────────────────────────────────────────────────────────────────
    uploaded          = request.files.get('file')
    dept_id           = request.form.get('department_id', type=int)
    auto_create_units = request.form.get('auto_create_units') == '1'

    def _err(msg, status=400):
        return jsonify({'error': msg}), status

    if not uploaded or not uploaded.filename:
        return _err('يرجى اختيار ملف')
    if not dept_id:
        return _err('يرجى اختيار القسم')

    filename = uploaded.filename.lower()
    rows = []

    if filename.endswith('.csv'):
        try:
            text = uploaded.read().decode('utf-8-sig')
            rows = list(csv.DictReader(io.StringIO(text)))
        except Exception as e:
            return _err(f'خطأ في قراءة ملف CSV: {e}')

    elif filename.endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded)
            ws = wb.active
            headers = [
                str(c.value).strip() if c.value else ''
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(
                    headers,
                    [str(v).strip() if v is not None else '' for v in row],
                )))
        except Exception as e:
            return _err(f'خطأ في قراءة ملف Excel: {e}')
    else:
        return _err('صيغة غير مدعومة — يُسمح بـ .xlsx و .csv فقط')

    # ── Lookup maps ──────────────────────────────────────────────────────────
    unit_map      = {u.name_ar.strip(): u for u in units}
    dept_map      = {d.name_ar.strip(): d for d in departments}
    dept_map_full = {
        f'{d.branch.name_ar.strip()} / {d.name_ar.strip()}': d
        for d in departments
    }
    cat_map = {c.name_ar.strip(): c for c in Category.query.all()}

    # ── Result buckets ───────────────────────────────────────────────────────
    created_items = []
    updated_items = []
    skipped_rows  = []
    warnings      = []
    units_created = 0

    # Within-file dedup guards
    seen_item_ids = set()          # raw item_id integers seen in this upload
    seen_codes    = set()          # (item_code, dept_id) pairs seen in this upload
    seen_pks      = set()          # DB item.id values already processed

    try:
        for row_num, row in enumerate(rows, start=2):
            # Skip fully empty rows
            if not any(str(v).strip() for v in row.values()):
                continue

            # ── Parse ─────────────────────────────────────────────────────────
            raw_id     = (row.get('item_id')       or '').strip()
            name_ar    = (row.get('item_name_ar')   or row.get('الصنف')              or row.get('name_ar')        or '').strip()
            name_en    = (row.get('item_name_en')   or row.get('name_en')            or '').strip() or None
            unit_name  = (row.get('unit_name')      or row.get('الوحدة')             or row.get('unit')           or '').strip()
            item_code  = (row.get('item_code')      or row.get('الكود')              or '').strip() or None
            pkg_note   = (row.get('notes')          or row.get('ملاحظة التغليف')     or row.get('packaging_note') or '').strip() or None
            cat_name   = (row.get('category_name')  or row.get('الفئة')              or row.get('category')       or '').strip() or None
            dept_name  = (row.get('القسم')          or row.get('department')         or '').strip() or None
            min_qty_s  = (
                row.get('min_level')    or row.get('minimum_stock') or
                row.get('min_stock')    or row.get('الحد الأدنى')  or ''
            ).strip()
            is_active_s = (row.get('is_active') or '').strip()

            # Extra conversion columns: الوحدة2/المعامل2 … الوحدة4/المعامل4
            extra_convs = []
            for n in ('2', '3', '4'):
                u_name = (row.get(f'الوحدة{n}') or '').strip()
                mult_s = (row.get(f'المعامل{n}') or '').strip()
                if u_name and mult_s:
                    extra_convs.append((u_name, mult_s, n))

            # ── Require at least a name or an id ─────────────────────────────
            if not name_ar and not raw_id:
                skipped_rows.append({'row': row_num, 'reason': 'الاسم العربي وitem_id كلاهما فارغان'})
                continue

            # ── Resolve department ────────────────────────────────────────────
            row_dept_id = dept_id
            if dept_name:
                d_obj = dept_map_full.get(dept_name) or dept_map.get(dept_name)
                if d_obj:
                    row_dept_id = d_obj.id
                else:
                    warnings.append(f'السطر {row_num}: قسم "{dept_name}" غير موجود — استخدام القسم الافتراضي')

            # ── Resolve unit ─────────────────────────────────────────────────
            unit_obj = unit_map.get(unit_name) if unit_name else None
            if unit_name and not unit_obj:
                if auto_create_units:
                    unit_obj = Unit(name_ar=unit_name, name_en=unit_name, symbol=unit_name, is_active=True)
                    db.session.add(unit_obj)
                    db.session.flush()
                    unit_map[unit_name] = unit_obj
                    units_created += 1
                    warnings.append(f'السطر {row_num}: وحدة "{unit_name}" أُنشئت تلقائياً')
                else:
                    skipped_rows.append({'row': row_num, 'name': name_ar, 'reason': f'وحدة "{unit_name}" غير موجودة'})
                    continue

            # ── Resolve category ─────────────────────────────────────────────
            cat_obj = cat_map.get(cat_name) if cat_name else None
            if cat_name and not cat_obj:
                warnings.append(f'السطر {row_num}: فئة "{cat_name}" غير موجودة — تم تجاهلها')

            # ── Parse min_qty ─────────────────────────────────────────────────
            min_qty = None
            if min_qty_s:
                try:
                    min_qty = float(min_qty_s)
                except ValueError:
                    warnings.append(f'السطر {row_num}: الحد الأدنى "{min_qty_s}" غير صالح — تم تجاهله')

            # ── Parse is_active ───────────────────────────────────────────────
            is_active_val = None
            if is_active_s:
                is_active_val = is_active_s.lower() in ('1', 'true', 'yes', 'نعم', 'active')

            # ── Validate extra conversion units ───────────────────────────────
            conv_pairs = []
            for u_name, mult_s, col_n in extra_convs:
                u_obj = unit_map.get(u_name)
                if not u_obj:
                    if auto_create_units:
                        u_obj = Unit(name_ar=u_name, name_en=u_name, symbol=u_name, is_active=True)
                        db.session.add(u_obj)
                        db.session.flush()
                        unit_map[u_name] = u_obj
                        units_created += 1
                        warnings.append(f'السطر {row_num}: وحدة{col_n} "{u_name}" أُنشئت تلقائياً')
                    else:
                        warnings.append(f'السطر {row_num}: وحدة{col_n} "{u_name}" غير موجودة — تم تجاهلها')
                        continue
                try:
                    mult = float(mult_s)
                    if mult <= 0:
                        raise ValueError
                except ValueError:
                    warnings.append(f'السطر {row_num}: معامل{col_n} "{mult_s}" غير صالح — تم تجاهله')
                    continue
                conv_pairs.append((u_obj, mult))

            # ── PRIORITY MATCHING ─────────────────────────────────────────────
            existing = None
            match_by = None

            # Step 1 — item_id (highest priority)
            if raw_id:
                try:
                    pk = int(float(raw_id))
                except (ValueError, TypeError):
                    warnings.append(f'السطر {row_num}: item_id "{raw_id}" غير صالح — تم تجاهله')
                    pk = None

                if pk is not None:
                    if pk in seen_item_ids:
                        skipped_rows.append({'row': row_num, 'reason': f'item_id {pk} مكرر في الملف'})
                        continue
                    seen_item_ids.add(pk)
                    candidate = db.session.get(Item, pk)
                    if candidate:
                        existing = candidate
                        match_by = 'item_id'
                    else:
                        warnings.append(
                            f'السطر {row_num}: item_id {pk} غير موجود في قاعدة البيانات'
                            ' — سيتم البحث بالكود أو الاسم'
                        )

            # Step 2 — item_code (within same department)
            if existing is None and item_code:
                if (item_code, row_dept_id) in seen_codes:
                    skipped_rows.append({
                        'row': row_num, 'name': name_ar,
                        'reason': f'كود "{item_code}" مكرر في الملف لنفس القسم',
                    })
                    continue
                seen_codes.add((item_code, row_dept_id))
                candidate = Item.query.filter_by(
                    item_code=item_code, department_id=row_dept_id
                ).first()
                if candidate:
                    existing = candidate
                    match_by = 'item_code'

            # Step 3 — name_ar, then name_en (within same department)
            if existing is None:
                if name_ar:
                    candidate = Item.query.filter_by(
                        name_ar=name_ar, department_id=row_dept_id
                    ).first()
                    if candidate:
                        existing = candidate
                        match_by = 'name_ar'
                if existing is None and name_en:
                    candidate = Item.query.filter_by(
                        name_en=name_en, department_id=row_dept_id
                    ).first()
                    if candidate:
                        existing = candidate
                        match_by = 'name_en'

            # Guard: don't touch the same DB row twice in one import
            if existing and existing.id in seen_pks:
                skipped_rows.append({
                    'row': row_num,
                    'name': name_ar or raw_id,
                    'reason': 'الصنف تمت معالجته مسبقاً في هذا الملف',
                })
                continue

            # ── UPDATE ────────────────────────────────────────────────────────
            if existing:
                changed = []

                if name_ar and name_ar != existing.name_ar:
                    existing.name_ar = name_ar
                    changed.append('الاسم العربي')
                if name_en:
                    existing.name_en = name_en
                    changed.append('الاسم الإنجليزي')
                if pkg_note is not None:
                    existing.packaging_note = pkg_note
                    changed.append('الملاحظات')
                if min_qty is not None:
                    existing.minimum_stock = min_qty
                    existing.min_quantity  = min_qty
                    changed.append('الحد الأدنى')
                if cat_obj:
                    existing.category_id = cat_obj.id
                    changed.append('الفئة')
                if is_active_val is not None:
                    existing.is_active = is_active_val
                    changed.append('الحالة')
                if unit_obj:
                    existing.unit_id      = unit_obj.id
                    existing.base_unit_id = unit_obj.id
                    changed.append('الوحدة الأساسية')
                    base_conv = ItemUnitConversion.query.filter_by(
                        item_id=existing.id, unit_id=unit_obj.id
                    ).first()
                    if not base_conv:
                        db.session.add(ItemUnitConversion(
                            item_id=existing.id, unit_id=unit_obj.id, multiplier=1.0
                        ))
                    elif base_conv.multiplier != 1.0:
                        base_conv.multiplier = 1.0

                # item_code: assign only if currently empty; never overwrite
                if item_code:
                    if not existing.item_code:
                        existing.item_code = item_code
                        changed.append('الكود')
                    elif existing.item_code != item_code:
                        warnings.append(
                            f'السطر {row_num}: كود "{item_code}" يختلف عن الكود الموجود'
                            f' "{existing.item_code}" — لم يتم التغيير'
                        )

                # Upsert extra unit conversions
                base_uid = existing.base_unit_id or existing.unit_id
                for u_obj, mult in conv_pairs:
                    if u_obj.id == base_uid:
                        continue
                    conv = ItemUnitConversion.query.filter_by(
                        item_id=existing.id, unit_id=u_obj.id
                    ).first()
                    if conv:
                        if conv.multiplier != mult:
                            conv.multiplier = mult
                            changed.append(f'معامل {u_obj.name_ar}')
                    else:
                        db.session.add(ItemUnitConversion(
                            item_id=existing.id, unit_id=u_obj.id, multiplier=mult
                        ))
                        changed.append(f'وحدة {u_obj.name_ar}')

                seen_pks.add(existing.id)
                updated_items.append({
                    'row':          row_num,
                    'id':           existing.id,
                    'name':         existing.name_ar,
                    'match_method': match_by,
                    'changes':      changed,
                })

            # ── CREATE ────────────────────────────────────────────────────────
            else:
                if not name_ar:
                    skipped_rows.append({'row': row_num, 'reason': 'اسم الصنف فارغ — لا يمكن الإنشاء'})
                    continue
                if not unit_obj:
                    skipped_rows.append({
                        'row': row_num, 'name': name_ar,
                        'reason': 'الوحدة مطلوبة لإنشاء صنف جديد',
                    })
                    continue

                new_item = Item(
                    item_code=item_code,
                    name_ar=name_ar,
                    name_en=name_en or '',
                    packaging_note=pkg_note,
                    unit_id=unit_obj.id,
                    base_unit_id=unit_obj.id,
                    department_id=row_dept_id,
                    min_quantity=min_qty or 0.0,
                    minimum_stock=min_qty or 0.0,
                    category_id=cat_obj.id if cat_obj else None,
                    is_active=is_active_val if is_active_val is not None else True,
                )
                db.session.add(new_item)
                db.session.flush()

                db.session.add(ItemUnitConversion(
                    item_id=new_item.id, unit_id=unit_obj.id, multiplier=1.0
                ))
                for u_obj, mult in conv_pairs:
                    if u_obj.id != unit_obj.id:
                        db.session.add(ItemUnitConversion(
                            item_id=new_item.id, unit_id=u_obj.id, multiplier=mult
                        ))

                seen_pks.add(new_item.id)
                created_items.append({
                    'row':  row_num,
                    'id':   new_item.id,
                    'name': new_item.name_ar,
                })

        db.session.commit()

    except Exception as exc:
        db.session.rollback()
        return jsonify({
            'error': f'خطأ حرج — تم التراجع عن جميع التغييرات: {exc}'
        }), 500

    return jsonify({
        'summary': {
            'created':       len(created_items),
            'updated':       len(updated_items),
            'skipped':       len(skipped_rows),
            'failed':        0,
            'units_created': units_created,
        },
        'created_items': created_items,
        'updated_items': updated_items,
        'skipped_rows':  skipped_rows,
        'warnings':      warnings,
    })
