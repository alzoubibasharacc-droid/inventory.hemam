from flask import Blueprint, render_template, request, redirect, url_for, flash, Response
from flask_login import login_required, current_user
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from models import db, Branch, Department, Item, InventoryCount, InventorySession, Unit, User
from datetime import datetime, date
import calendar
import io
from utils.constants import now_jordan

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

# Excel brand colours (used only in _build_xlsx)
_GREEN  = '295831'
_OLIVE  = '4B5947'
_ORANGE = 'F28705'
_WHITE  = 'FFFFFF'


# ── Query helpers ─────────────────────────────────────────────────────────────

def _base_filters(q, branch_id, dept_id, user_id, date_from, date_to, session_id=None):
    """
    Apply shared filter conditions to an InventoryCount query.

    Primary:  session_id (when provided — session is the new grouping key)
    Fallback: count_date range (backward-compat date-based filtering)
    """
    if session_id:
        q = q.filter(InventoryCount.session_id == session_id)
    else:
        if date_from:
            q = q.filter(InventoryCount.count_date >= date_from)
        if date_to:
            q = q.filter(InventoryCount.count_date <= date_to)
    if branch_id:
        q = q.filter(Department.branch_id == branch_id)
    if dept_id:
        q = q.filter(Item.department_id == dept_id)
    if user_id:
        q = q.filter(InventoryCount.user_id == user_id)
    return q


def _build_entry_query(branch_id, dept_id, user_id, date_from, date_to, session_id=None):
    """Return a fully-filtered InventoryCount query with eager-loaded relationships."""
    q = (
        InventoryCount.query
        .join(Item, InventoryCount.item_id == Item.id)
        .join(Department, Item.department_id == Department.id)
        .options(
            joinedload(InventoryCount.item).options(
                joinedload(Item.department).joinedload(Department.branch),
                joinedload(Item.base_unit),
                joinedload(Item.unit),
            ),
            joinedload(InventoryCount.user),
            joinedload(InventoryCount.entered_unit),
        )
    )
    return _base_filters(q, branch_id, dept_id, user_id, date_from, date_to, session_id)


def _item_totals_sql(branch_id, dept_id, user_id, date_from, date_to, session_id=None):
    """Return a dict {item_id: total_base_quantity} computed entirely in SQL."""
    q = (
        db.session.query(
            InventoryCount.item_id,
            func.sum(InventoryCount.quantity).label('total'),
        )
        .join(Item, InventoryCount.item_id == Item.id)
        .join(Department, Item.department_id == Department.id)
        .group_by(InventoryCount.item_id)
    )
    q = _base_filters(q, branch_id, dept_id, user_id, date_from, date_to, session_id)
    return {row.item_id: (row.total or 0.0) for row in q.all()}


def _parse_date_range():
    """Parse date_from / date_to from request args. Defaults to current month."""
    now = now_jordan()
    date_from = date_to = None
    try:
        df_raw = request.args.get('date_from', '').strip()
        date_from = datetime.strptime(df_raw, '%Y-%m-%d').date() if df_raw else None
    except ValueError:
        pass
    try:
        dt_raw = request.args.get('date_to', '').strip()
        date_to = datetime.strptime(dt_raw, '%Y-%m-%d').date() if dt_raw else None
    except ValueError:
        pass
    if not date_from and not date_to:
        date_from = date(now.year, now.month, 1)
        date_to   = date(now.year, now.month, calendar.monthrange(now.year, now.month)[1])
    return date_from, date_to


def _parse_export_filters():
    """Parse all export filters from request args."""
    branch_id  = request.args.get('branch_id',  type=int)
    dept_id    = request.args.get('dept_id',     type=int)
    user_id    = request.args.get('user_id',     type=int)
    session_id = request.args.get('session_id',  type=int)
    date_from, date_to = _parse_date_range()
    return branch_id, dept_id, user_id, date_from, date_to, session_id


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_xlsx(counts, filters_label):
    """Build and return openpyxl Workbook bytes for the given InventoryCount list."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير الجرد'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill('solid', fgColor=_OLIVE)
    alt_fill    = PatternFill('solid', fgColor='EEF0EE')
    total_fill  = PatternFill('solid', fgColor='D4E6D6')

    header_font = Font(name='Cairo', bold=True,   color=_WHITE,  size=11)
    title_font  = Font(name='Cairo', bold=True,   color=_OLIVE,  size=14)
    sub_font    = Font(name='Cairo', italic=True, color='666666', size=10)
    data_font   = Font(name='Cairo', size=10)
    total_font  = Font(name='Cairo', bold=True,   color=_GREEN,  size=11)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
    right  = Alignment(horizontal='right',  vertical='center', wrap_text=True, readingOrder=2)

    thin  = Side(style='thin',   color='CCCCCC')
    thick = Side(style='medium', color=_OLIVE)
    cell_border   = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    header_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    COLS = 11

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
    t = ws.cell(1, 1, 'تقرير الجرد - مطاعم همم')
    t.font, t.alignment = title_font, center
    t.fill = PatternFill('solid', fgColor='F4F6F4')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COLS)
    s = ws.cell(2, 1, filters_label)
    s.font, s.alignment = sub_font, center

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    headers = [
        'كود الصنف', 'اسم الصنف', 'ملاحظة التغليف', 'الفرع', 'القسم',
        'الكمية المُدخَلة', 'وحدة الإدخال', 'الكمية (وحدة أساسية)',
        'الموظف', 'تاريخ ووقت الجرد', 'ملاحظات',
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(3, col, h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, header_border
    ws.row_dimensions[3].height = 24

    total_base = 0.0
    for row_i, cnt in enumerate(counts, start=4):
        fill = PatternFill('solid', fgColor=_WHITE) if row_i % 2 == 0 else alt_fill
        entered_qty  = cnt.entered_quantity if cnt.entered_quantity is not None else cnt.quantity
        entered_unit = (
            cnt.entered_unit.name_ar if cnt.entered_unit
            else cnt.item.effective_base_unit.name_ar
        )
        row_data = [
            cnt.item.item_code or '',
            cnt.item.name_ar,
            cnt.item.packaging_note or '',
            cnt.item.department.branch.name_ar,
            cnt.item.department.name_ar,
            entered_qty,
            entered_unit,
            cnt.quantity,
            cnt.user.full_name,
            cnt.created_at.strftime('%Y-%m-%d  %H:%M') if cnt.created_at else '',
            cnt.notes or '',
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row_i, col, val)
            c.font, c.fill, c.border = data_font, fill, cell_border
            c.alignment = center if col in (6, 8) else right
            if col in (6, 8):
                c.number_format = '#,##0.###'
        total_base += cnt.quantity
        ws.row_dimensions[row_i].height = 18

    total_row = 4 + len(counts)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    label_c = ws.cell(total_row, 1, f'الإجمالي  ({len(counts)} إدخال)')
    label_c.font, label_c.fill, label_c.alignment, label_c.border = (
        total_font, total_fill, center, cell_border
    )
    qty_c = ws.cell(total_row, 8, total_base)
    qty_c.font, qty_c.fill, qty_c.alignment = total_font, total_fill, center
    qty_c.number_format, qty_c.border = '#,##0.###', cell_border
    for col in range(9, COLS + 1):
        c = ws.cell(total_row, col)
        c.fill, c.border = total_fill, cell_border
    ws.row_dimensions[total_row].height = 22

    col_min_widths = [14, 28, 22, 18, 16, 14, 14, 18, 18, 22, 30]
    for col_i, min_w in enumerate(col_min_widths, start=1):
        max_len = min_w
        for row in ws.iter_rows(min_row=3, min_col=col_i, max_col=col_i):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) * 1.4)
        ws.column_dimensions[get_column_letter(col_i)].width = min(max_len, 45)

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Routes ────────────────────────────────────────────────────────────────────

@reports_bp.route('/')
@login_required
def index():
    now = now_jordan()

    date_from_str = request.args.get('date_from',  '')
    date_to_str   = request.args.get('date_to',    '')
    branch_id     = request.args.get('branch_id',  type=int)
    dept_id       = request.args.get('dept_id',    type=int)
    session_id    = request.args.get('session_id', type=int)

    # Block employees from manually accessing outside their assigned scope
    if not current_user.is_manager:
        if current_user.branch_id and branch_id and branch_id != current_user.branch_id:
            flash('وصول غير مصرح به', 'danger')
            return redirect(url_for('inventory.dashboard'))
        if current_user.department_id and dept_id and dept_id != current_user.department_id:
            flash('وصول غير مصرح به', 'danger')
            return redirect(url_for('inventory.dashboard'))

    date_from = date_to = None
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    except ValueError:
        pass
    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        pass

    # Default to current month
    if not date_from and not date_to:
        date_from = date(now.year, now.month, 1)
        date_to   = date(now.year, now.month, calendar.monthrange(now.year, now.month)[1])
        date_from_str = date_from.isoformat()
        date_to_str   = date_to.isoformat()

    # Enforce branch + department scope for non-admins
    if not current_user.is_admin and current_user.branch_id:
        branch_id = current_user.branch_id
    if not current_user.is_manager and current_user.department_id:
        dept_id = current_user.department_id

    if not current_user.is_manager and current_user.branch_id:
        branches    = [current_user.branch]
        departments = (
            [Department.query.get(current_user.department_id)]
            if current_user.department_id else []
        )
    else:
        branches    = Branch.query.all()
        departments = Department.query.filter_by(branch_id=branch_id).all() if branch_id else []

    # Resolve session object for display (None when filtering by date range)
    active_session = InventorySession.query.get(session_id) if session_id else None

    totals = _item_totals_sql(branch_id, dept_id, None, date_from, date_to, session_id)

    all_entries = (
        _build_entry_query(branch_id, dept_id, None, date_from, date_to, session_id)
        .order_by(Department.name_ar, Item.name_ar, InventoryCount.created_at)
        .all()
    )

    dept_order = []
    item_agg   = {}
    for entry in all_entries:
        dept_name = entry.item.department.name_ar
        if dept_name not in dept_order:
            dept_order.append(dept_name)
        iid = entry.item_id
        if iid not in item_agg:
            tb = totals.get(iid, 0.0)
            ms = entry.item.effective_minimum_stock
            item_agg[iid] = {
                'item':           entry.item,
                'total_base':     tb,
                'entries':        [],
                'dept':           dept_name,
                'minimum_stock':  ms,
                'low_stock':      ms > 0 and tb < ms,
                'is_counted':     True,
            }
        item_agg[iid]['entries'].append(entry)

    grouped = {d: [] for d in dept_order}
    for agg in item_agg.values():
        grouped[agg['dept']].append(agg)

    # Append uncounted items when a session is selected
    if session_id and active_session:
        dept_ids_in_scope = [
            sd.department_id for sd in active_session.session_departments
        ]
        if dept_id:
            dept_ids_in_scope = [d for d in dept_ids_in_scope if d == dept_id]
        counted_ids = set(item_agg.keys())
        if dept_ids_in_scope:
            q_uncounted = (
                db.session.query(Item, Department)
                .join(Department, Item.department_id == Department.id)
                .filter(
                    Item.department_id.in_(dept_ids_in_scope),
                    Item.is_active == True,
                )
                .order_by(Department.name_ar, Item.name_ar)
            )
            if counted_ids:
                q_uncounted = q_uncounted.filter(~Item.id.in_(counted_ids))
            for item, dept in q_uncounted.all():
                dept_name = dept.name_ar
                if dept_name not in grouped:
                    grouped[dept_name] = []
                grouped[dept_name].append({
                    'item':          item,
                    'total_base':    0,
                    'entries':       [],
                    'dept':          dept_name,
                    'minimum_stock': item.effective_minimum_stock,
                    'low_stock':     False,
                    'is_counted':    False,
                })

    # Build inv_sessions list for the filter dropdown (scoped for employees)
    inv_sessions_q = (
        InventorySession.query
        .options(joinedload(InventorySession.branch))
        .filter(InventorySession.session_type != 'baseline')
        .order_by(InventorySession.count_date.desc(), InventorySession.branch_id)
    )
    if not current_user.is_manager and current_user.branch_id:
        inv_sessions_q = inv_sessions_q.filter(
            InventorySession.branch_id == current_user.branch_id
        )
    inv_sessions = inv_sessions_q.all()

    return render_template(
        'inventory/reports.html',
        grouped=grouped,
        branches=branches,
        departments=departments,
        inv_sessions=inv_sessions,
        selected_branch=branch_id,
        selected_dept=dept_id,
        selected_session=session_id,
        active_session=active_session,
        date_from=date_from_str,
        date_to=date_to_str,
        now=now,
    )


@reports_bp.route('/export')
@login_required
def export_page():
    if not current_user.is_admin:
        flash('صلاحية المدير مطلوبة لتصدير البيانات', 'danger')
        return redirect(url_for('reports.index'))

    now         = now_jordan()
    branches    = Branch.query.all()
    departments = Department.query.all()
    employees   = User.query.filter_by(is_active=True).order_by(User.full_name).all()

    branch_id     = request.args.get('branch_id',  type=int)
    dept_id       = request.args.get('dept_id',    type=int)
    user_id       = request.args.get('user_id',    type=int)
    session_id    = request.args.get('session_id', type=int)
    date_from_str = request.args.get('date_from',  '')
    date_to_str   = request.args.get('date_to',    '')

    date_from = date_to = None
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    except ValueError:
        pass
    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        pass

    preview_count = None
    if request.args:
        preview_count = (
            _build_entry_query(branch_id, dept_id, user_id, date_from, date_to, session_id)
            .count()
        )

    inv_sessions = (
        InventorySession.query
        .options(joinedload(InventorySession.branch))
        .order_by(InventorySession.branch_id, InventorySession.count_date.desc())
        .all()
    )

    return render_template(
        'inventory/export.html',
        branches=branches,
        departments=departments,
        employees=employees,
        inv_sessions=inv_sessions,
        now=now,
        selected_branch=branch_id,
        selected_dept=dept_id,
        selected_user=user_id,
        selected_session=session_id,
        date_from=date_from_str,
        date_to=date_to_str,
        preview_count=preview_count,
    )


@reports_bp.route('/export/download')
@login_required
def export_download():
    if not current_user.is_admin:
        flash('صلاحية المدير مطلوبة لتصدير البيانات', 'danger')
        return redirect(url_for('reports.index'))

    branch_id, dept_id, user_id, date_from, date_to, session_id = _parse_export_filters()

    counts = (
        _build_entry_query(branch_id, dept_id, user_id, date_from, date_to, session_id)
        .order_by(Department.branch_id, Department.name_ar, Item.name_ar)
        .all()
    )

    if not counts:
        flash('لا توجد بيانات مطابقة للفلتر المحدد', 'warning')
        return redirect(url_for('reports.export_page', **request.args))

    active_session = InventorySession.query.get(session_id) if session_id else None

    parts = []
    if active_session:
        parts.append(f'{active_session.name} | {active_session.count_date}')
        if active_session.branch:
            parts.append(active_session.branch.name_ar)
    else:
        if date_from and date_to:
            parts.append(f'من {date_from} إلى {date_to}')
        elif date_from:
            parts.append(f'من {date_from}')
        elif date_to:
            parts.append(f'حتى {date_to}')
        if branch_id:
            b = Branch.query.get(branch_id)
            if b: parts.append(b.name_ar)

    if dept_id:
        d = Department.query.get(dept_id)
        if d: parts.append(d.name_ar)
    if user_id:
        u = User.query.get(user_id)
        if u: parts.append(u.full_name)

    filters_label = '  |  '.join(parts) if parts else 'كل الفترات'

    if active_session:
        safe_name = ''.join(c for c in active_session.name if c.isalnum() or c in ' _-').strip()
        filename = f'تقرير_جرد_{safe_name}_{active_session.count_date}.xlsx'
    else:
        filename = f'inventory_{date_from}_{date_to}.xlsx'

    xlsx_bytes = _build_xlsx(counts, filters_label)
    return Response(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@reports_bp.route('/converter')
@login_required
def converter():
    units = Unit.query.order_by(Unit.name_ar).all()
    return render_template('inventory/converter.html', units=units)
