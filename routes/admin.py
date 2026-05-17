from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from models import db, Item, Unit, UnitConversion, Department, Branch, User, Category, ItemUnitConversion
from functools import wraps
import csv
import io

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


# ── decorators ───────────────────────────────────────────────────────────────

def manager_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_manager:
            flash('ليس لديك صلاحية الوصول', 'danger')
            return redirect(url_for('inventory.dashboard'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            flash('هذه الصفحة للمدير فقط', 'danger')
            return redirect(url_for('inventory.dashboard'))
        return f(*args, **kwargs)
    return decorated


# ── items ────────────────────────────────────────────────────────────────────

@admin_bp.route('/items', methods=['GET', 'POST'])
@manager_required
def items():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            code = request.form.get('item_code', '').strip() or None
            if code and Item.query.filter_by(item_code=code).first():
                flash(f'كود الصنف "{code}" مستخدم مسبقاً', 'danger')
                return redirect(url_for('admin.items'))

            base_unit_id = request.form.get('base_unit_id') or request.form.get('unit_id')
            item = Item(
                item_code=code,
                name_ar=request.form['name_ar'].strip(),
                name_en=request.form.get('name_en', '').strip(),
                packaging_note=request.form.get('packaging_note', '').strip() or None,
                unit_id=int(request.form['unit_id']),
                base_unit_id=int(base_unit_id),
                department_id=int(request.form['department_id']),
                category_id=request.form.get('category_id') or None,
                min_quantity=float(request.form.get('min_quantity', 0) or 0)
            )
            db.session.add(item)
            db.session.flush()  # get item.id before adding conversion

            # Auto-create base-unit conversion (multiplier = 1)
            base_conv = ItemUnitConversion(
                item_id=item.id,
                unit_id=item.base_unit_id,
                multiplier=1.0
            )
            db.session.add(base_conv)
            db.session.commit()
            flash('تم إضافة الصنف بنجاح', 'success')

        elif action == 'toggle':
            item = Item.query.get_or_404(int(request.form['item_id']))
            item.is_active = not item.is_active
            db.session.commit()
            flash('تم تحديث حالة الصنف', 'info')

        return redirect(url_for('admin.items'))

    branch_filter = request.args.get('branch_id', type=int)
    dept_filter   = request.args.get('dept_id',   type=int)
    search_q      = request.args.get('q', '').strip()

    items_q = Item.query.join(Department)
    if branch_filter:
        items_q = items_q.filter(Department.branch_id == branch_filter)
    if dept_filter:
        items_q = items_q.filter(Item.department_id == dept_filter)
    if search_q:
        like = f'%{search_q}%'
        items_q = items_q.filter(
            db.or_(Item.name_ar.ilike(like), Item.item_code.ilike(like))
        )
    items_list = items_q.order_by(Department.name_ar, Item.name_ar).all()

    branches    = Branch.query.all()
    departments = Department.query.all()
    units       = Unit.query.order_by(Unit.name_ar).all()
    categories  = Category.query.all()

    return render_template('admin/items.html',
                           items=items_list,
                           branches=branches, departments=departments,
                           units=units, categories=categories,
                           selected_branch=branch_filter,
                           selected_dept=dept_filter,
                           search_q=search_q)


# ── item unit conversions ────────────────────────────────────────────────────

@admin_bp.route('/items/<int:item_id>/conversions', methods=['GET', 'POST'])
@manager_required
def item_conversions(item_id):
    item = Item.query.get_or_404(item_id)
    units = Unit.query.order_by(Unit.name_ar).all()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            unit_id   = int(request.form['unit_id'])
            try:
                multiplier = float(request.form['multiplier'])
            except ValueError:
                flash('المعامل يجب أن يكون رقماً', 'danger')
                return redirect(url_for('admin.item_conversions', item_id=item_id))

            if multiplier <= 0:
                flash('المعامل يجب أن يكون أكبر من صفر', 'danger')
                return redirect(url_for('admin.item_conversions', item_id=item_id))

            existing = ItemUnitConversion.query.filter_by(
                item_id=item_id, unit_id=unit_id
            ).first()
            if existing:
                existing.multiplier = multiplier
                flash('تم تحديث معامل التحويل', 'info')
            else:
                conv = ItemUnitConversion(
                    item_id=item_id, unit_id=unit_id, multiplier=multiplier
                )
                db.session.add(conv)
                flash('تم إضافة وحدة التحويل', 'success')
            db.session.commit()

        elif action == 'delete':
            conv_id = int(request.form['conv_id'])
            conv = ItemUnitConversion.query.get_or_404(conv_id)
            # Protect base unit: cannot delete multiplier=1 if it's the only conversion
            if conv.unit_id == item.effective_base_unit_id and conv.multiplier == 1.0:
                count = ItemUnitConversion.query.filter_by(item_id=item_id).count()
                if count <= 1:
                    flash('لا يمكن حذف الوحدة الأساسية', 'danger')
                    return redirect(url_for('admin.item_conversions', item_id=item_id))
            db.session.delete(conv)
            db.session.commit()
            flash('تم حذف وحدة التحويل', 'info')

        return redirect(url_for('admin.item_conversions', item_id=item_id))

    conversions = ItemUnitConversion.query.filter_by(item_id=item_id)\
                      .order_by(ItemUnitConversion.multiplier).all()

    return render_template('admin/item_conversions.html',
                           item=item, conversions=conversions, units=units)


# ── users ────────────────────────────────────────────────────────────────────

@admin_bp.route('/users', methods=['GET', 'POST'])
@admin_required
def users():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            if User.query.filter_by(username=request.form['username']).first():
                flash('اسم المستخدم موجود مسبقاً', 'danger')
            else:
                user = User(
                    username=request.form['username'].strip(),
                    full_name=request.form['full_name'].strip(),
                    role=request.form['role'],
                    branch_id=request.form.get('branch_id') or None,
                    department_id=request.form.get('department_id') or None
                )
                user.set_password(request.form['password'])
                db.session.add(user)
                db.session.commit()
                flash('تم إضافة المستخدم بنجاح', 'success')
        elif action == 'toggle':
            user = User.query.get_or_404(int(request.form['user_id']))
            if user.id != current_user.id:
                user.is_active = not user.is_active
                db.session.commit()
                flash('تم تحديث حالة المستخدم', 'info')
        elif action == 'reset_password':
            user = User.query.get_or_404(int(request.form['user_id']))
            user.set_password(request.form['new_password'])
            db.session.commit()
            flash('تم تغيير كلمة المرور', 'success')
        return redirect(url_for('admin.users'))

    users_list  = User.query.order_by(User.full_name).all()
    branches    = Branch.query.all()
    departments = Department.query.all()
    return render_template('admin/users.html',
                           users=users_list,
                           branches=branches, departments=departments)


# ── global units ─────────────────────────────────────────────────────────────

@admin_bp.route('/units', methods=['GET', 'POST'])
@admin_required
def units():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add_unit':
            unit = Unit(
                name_ar=request.form['name_ar'].strip(),
                name_en=request.form.get('name_en', '').strip()
            )
            db.session.add(unit)
            db.session.commit()
            flash('تم إضافة وحدة القياس', 'success')
        elif action == 'add_conversion':
            conv = UnitConversion(
                from_unit_id=int(request.form['from_unit_id']),
                to_unit_id=int(request.form['to_unit_id']),
                factor=float(request.form['factor'])
            )
            db.session.add(conv)
            db.session.commit()
            flash('تم إضافة التحويل', 'success')
        return redirect(url_for('admin.units'))

    units_list  = Unit.query.order_by(Unit.name_ar).all()
    conversions = UnitConversion.query.all()
    return render_template('admin/units.html',
                           units=units_list, conversions=conversions)


# ── excel import ─────────────────────────────────────────────────────────────

@admin_bp.route('/import', methods=['GET', 'POST'])
@manager_required
def import_items():
    branches    = Branch.query.all()
    departments = Department.query.all()
    units       = Unit.query.order_by(Unit.name_ar).all()

    if request.method == 'POST':
        uploaded = request.files.get('file')
        dept_id  = request.form.get('department_id', type=int)

        if not uploaded or uploaded.filename == '':
            flash('يرجى اختيار ملف', 'warning')
            return redirect(url_for('admin.import_items'))
        if not dept_id:
            flash('يرجى اختيار القسم', 'warning')
            return redirect(url_for('admin.import_items'))

        filename = uploaded.filename.lower()
        rows, errors = [], []

        if filename.endswith('.csv'):
            try:
                text = uploaded.read().decode('utf-8-sig')
                rows = list(csv.DictReader(io.StringIO(text)))
            except Exception as e:
                flash(f'خطأ في قراءة ملف CSV: {e}', 'danger')
                return redirect(url_for('admin.import_items'))
        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(uploaded)
                ws = wb.active
                headers = [
                    str(c.value).strip() if c.value else ''
                    for c in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(
                        headers,
                        [str(v).strip() if v is not None else '' for v in row]
                    )))
            except Exception as e:
                flash(f'خطأ في قراءة ملف Excel: {e}', 'danger')
                return redirect(url_for('admin.import_items'))
        else:
            flash('صيغة غير مدعومة — يُسمح بـ .xlsx و .csv فقط', 'danger')
            return redirect(url_for('admin.import_items'))

        unit_map = {u.name_ar.strip(): u for u in units}
        imported = skipped = 0

        # Collect item codes seen within this file to catch intra-file duplicates
        file_codes_seen = set()

        for i, row in enumerate(rows, start=2):
            name_ar   = (row.get('الصنف')         or row.get('name_ar')        or '').strip()
            unit_name = (row.get('الوحدة')         or row.get('unit')           or '').strip()
            min_qty_s = (row.get('الحد الأدنى')   or row.get('min_quantity')   or '0').strip()
            item_code = (row.get('الكود')           or row.get('item_code')      or '').strip() or None
            pkg_note  = (row.get('ملاحظة التغليف') or row.get('packaging_note') or '').strip() or None
            name_en   = (row.get('name_en')                                      or '').strip()

            # Extra conversion columns: الوحدة2/المعامل2, الوحدة3/المعامل3
            extra_convs = []
            for n in ('2', '3', '4'):
                u_name = (row.get(f'الوحدة{n}') or '').strip()
                mult_s = (row.get(f'المعامل{n}') or '').strip()
                if u_name and mult_s:
                    extra_convs.append((u_name, mult_s, n))

            # ── Validation ────────────────────────────────────────────────────
            if not name_ar:
                errors.append(f'السطر {i}: اسم الصنف فارغ — تخطي')
                skipped += 1
                continue

            unit_obj = unit_map.get(unit_name)
            if not unit_obj:
                errors.append(f'السطر {i}: وحدة "{unit_name}" غير موجودة — تخطي')
                skipped += 1
                continue

            try:
                min_qty = float(min_qty_s or '0')
            except ValueError:
                min_qty = 0.0

            # item_code: intra-file duplicate check
            if item_code:
                if item_code in file_codes_seen:
                    errors.append(f'السطر {i}: كود "{item_code}" مكرر في الملف — تخطي')
                    skipped += 1
                    continue
                if Item.query.filter_by(item_code=item_code).first():
                    errors.append(f'السطر {i}: كود "{item_code}" موجود مسبقاً في قاعدة البيانات — تخطي')
                    skipped += 1
                    continue
                file_codes_seen.add(item_code)

            if Item.query.filter_by(name_ar=name_ar, department_id=dept_id).first():
                errors.append(f'السطر {i}: "{name_ar}" موجود مسبقاً في هذا القسم — تخطي')
                skipped += 1
                continue

            # Validate extra conversion units before inserting
            conv_pairs = []
            conv_ok = True
            for u_name, mult_s, col_n in extra_convs:
                u_obj = unit_map.get(u_name)
                if not u_obj:
                    errors.append(f'السطر {i}: وحدة{col_n} "{u_name}" غير موجودة — تم إضافة الصنف بدون هذه الوحدة')
                    continue
                try:
                    mult = float(mult_s)
                    if mult <= 0:
                        raise ValueError
                except ValueError:
                    errors.append(f'السطر {i}: معامل{col_n} "{mult_s}" غير صالح — تم إضافة الصنف بدون هذه الوحدة')
                    continue
                conv_pairs.append((u_obj, mult))

            # ── Insert ────────────────────────────────────────────────────────
            item = Item(
                item_code=item_code,
                name_ar=name_ar,
                name_en=name_en,
                packaging_note=pkg_note,
                unit_id=unit_obj.id,
                base_unit_id=unit_obj.id,
                department_id=dept_id,
                min_quantity=min_qty,
            )
            db.session.add(item)
            db.session.flush()

            # Base unit conversion (multiplier = 1)
            db.session.add(
                ItemUnitConversion(item_id=item.id, unit_id=unit_obj.id, multiplier=1.0)
            )

            # Extra conversions
            for u_obj, mult in conv_pairs:
                if u_obj.id != unit_obj.id:  # skip if same as base (already added)
                    db.session.add(
                        ItemUnitConversion(item_id=item.id, unit_id=u_obj.id, multiplier=mult)
                    )

            imported += 1

        db.session.commit()
        msg = f'تم استيراد {imported} صنف' + (f' (تخطي {skipped})' if skipped else '')
        flash(msg, 'success' if imported else 'warning')
        return render_template('admin/import.html',
                               branches=branches, departments=departments, units=units,
                               errors=errors, imported=imported, skipped=skipped)

    return render_template('admin/import.html',
                           branches=branches, departments=departments, units=units,
                           errors=None, imported=None, skipped=None)
